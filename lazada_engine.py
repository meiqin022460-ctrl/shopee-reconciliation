import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, warnings

warnings.filterwarnings('ignore')

C_BLUE   = 'ADD8E6'
C_ORANGE = 'FFD580'
C_RED    = 'FFB3B3'
C_YELLOW = 'FFFAAA'
C_PINK   = 'F7B2E3'
C_GRAY   = 'D3D3D3'
C_GREEN  = 'C8E6C9'
C_HEADER = '2F4F8F'
C_WHITE  = 'FFFFFF'

FEE_COLS = ['Payment Fee', 'Commission', 'LazCoins Discount',
            'LazCoins Discount Promotion Fee', 'Sponsored Affiliates']

CATEGORY_MAP = [
    ('Sales', ['Item Price Credit']),
    ('Lazada Fees', ['Commission', 'Payment Fee', 'LazCoins Discount Promotion Fee',
                      'Payment fee - correction for undercharge',
                      'Commission fee - correction for undercharge',
                      'Buyer Review Incentive']),
    ('Marketing Fees', ['LazCoins Discount', 'Sponsored Affiliates']),
    ('Logistics', []),
    ('Claims', ['Lost Claim']),
]


def mk_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def mk_border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)


def write_note(ws, text, n_cols, height=44):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(italic=True, size=9, color='555555')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = height


def write_header(ws, row_num, labels, widths=None, bg=C_HEADER):
    f = mk_fill(bg)
    ft = Font(bold=True, color=C_WHITE, size=10)
    al = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, lbl in enumerate(labels, 1):
        c = ws.cell(row=row_num, column=i, value=lbl)
        c.fill = f; c.font = ft; c.alignment = al; c.border = mk_border()
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, df, col_order, amt_cols, start_row=2, default_color='FFFFFF', color_col='row_color', pct_cols=None):
    pct_cols = pct_cols or set()
    for r_i, (_, row) in enumerate(df.iterrows(), start_row):
        clr = row.get(color_col, default_color) if color_col in df.columns else default_color
        for c_i, col in enumerate(col_order, 1):
            val = row.get(col, '')
            if isinstance(val, float) and pd.isna(val):
                val = ''
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.fill = mk_fill(clr); cell.border = mk_border()
            if col in pct_cols:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in amt_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in ('Order Number', 'Inv No'):
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')


def _read_many(files, **kwargs):
    frames = []
    for f in files:
        d = pd.read_excel(f, **kwargs)
        d.columns = [str(c).strip() for c in d.columns]
        frames.append(d)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def build_lazada_report(income_files, accinv_files, or_files, order_files, wallet_files,
                         filter_start, filter_end):
    """Build the full Lazada reconciliation workbook from uploaded files.

    income_files / accinv_files: required, list of uploaded file objects.
    or_files / order_files / wallet_files: optional, list of uploaded file objects (may be empty).
    filter_start / filter_end: pd.Timestamp, the month/period to report on.

    Returns (buf, stats) where buf is a io.BytesIO holding the .xlsx file
    and stats is a dict of headline numbers for display.
    """
    FILTER_START, FILTER_END = filter_start, filter_end

    # ── Load Income Overview ───────────────────────────
    inc_all = _read_many(income_files, sheet_name='Income Overview',
                          dtype={'Order Number': str, 'Order Line ID': str})
    inc_all['Amount(Include Tax)'] = pd.to_numeric(inc_all['Amount(Include Tax)'], errors='coerce').fillna(0)
    inc_all['Transaction Date_dt'] = pd.to_datetime(inc_all['Transaction Date'], format='%d %b %Y', errors='coerce')
    inc_all['Order Creation Date_dt'] = pd.to_datetime(inc_all['Order Creation Date'], format='%d %b %Y', errors='coerce')
    inc_all = inc_all.drop_duplicates()

    orders_in_period = set(inc_all.loc[
        (inc_all['Order Creation Date_dt'] >= FILTER_START) &
        (inc_all['Order Creation Date_dt'] <= FILTER_END), 'Order Number'])
    inc = inc_all[inc_all['Order Number'].isin(orders_in_period)].copy()

    orders_by_settlement = set(inc_all.loc[
        (inc_all['Transaction Date_dt'] >= FILTER_START) &
        (inc_all['Transaction Date_dt'] <= FILTER_END), 'Order Number'])
    inc_settlement = inc_all[inc_all['Order Number'].isin(orders_by_settlement)].copy()

    inc_period = inc_all[(inc_all['Transaction Date_dt'] >= FILTER_START) &
                          (inc_all['Transaction Date_dt'] <= FILTER_END)].copy()

    # ── Load ACC INV ────────────────────────────────────
    acc = _read_many(accinv_files, sheet_name='Sheet')
    acc['Order Number'] = acc['Shipping Info'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    acc = acc[acc['Order Number'] != ''].drop_duplicates(subset=['Doc. No.'])
    acc_full = acc.copy()
    acc_in_period = acc[(acc['Date'] >= FILTER_START) & (acc['Date'] <= FILTER_END)].copy()
    acc_info = acc_full.set_index('Order Number').to_dict('index')

    # ── Load OR (collection records) ───────────────────
    df_or = _read_many(or_files, sheet_name=0) if or_files else pd.DataFrame(
        columns=['Type', 'Date', 'No.', 'Org. Amt.', 'Outstanding', 'Paid Amount'])
    if 'No.' in df_or.columns:
        df_or['No.'] = df_or['No.'].astype(str).str.strip()
    or_map = {}
    for _, row in df_or.iterrows():
        no = str(row.get('No.', '')).strip()
        if not no or no == 'nan':
            continue
        org = float(row.get('Org. Amt.', 0) or 0)
        paid = float(row.get('Paid Amount', 0) or 0)
        outstanding = float(row.get('Outstanding', 0) or 0)
        if no in or_map:
            or_map[no]['org_amt'] += org
            or_map[no]['paid'] += paid
            or_map[no]['outstanding'] += outstanding
        else:
            or_map[no] = {'org_amt': org, 'paid': paid, 'outstanding': outstanding}

    # ── Load Order Export (real Shipping Fee source) ───
    real_shipfee_map = {}
    real_shipfee_by_line = {}
    if order_files:
        orders_export = _read_many(order_files, sheet_name=0,
                                    dtype={'orderNumber': str, 'orderItemId': str})
        orders_export = orders_export.drop_duplicates(subset=['orderItemId'])
        real_shipfee_map = orders_export.groupby('orderNumber')['shippingFee'].sum(min_count=1).to_dict()
        real_shipfee_by_line = orders_export.set_index('orderItemId')['shippingFee'].to_dict()

    # ── Load Wallet (Balance Transactions) ─────────────
    df_wallet = pd.DataFrame()
    if wallet_files:
        df_wallet = _read_many(wallet_files, sheet_name='Balance Transactions')
        df_wallet['Amount'] = (df_wallet['Amount'].astype(str)
                                .str.replace(',', '').str.replace('+', '').astype(float))
        df_wallet['Statement No'] = df_wallet['Remarks'].astype(str).str.extract(r'(MY\S+-\d{4}-\d{4})')
        stmt_digits = df_wallet['Remarks'].astype(str).str.extract(r'(\d{4}-?\d{4})')[0].str.replace('-', '', regex=False)
        df_wallet['Statement Date_dt'] = pd.to_datetime(stmt_digits, format='%Y%m%d', errors='coerce')
        df_wallet = df_wallet[(df_wallet['Statement Date_dt'] >= FILTER_START) &
                               (df_wallet['Statement Date_dt'] <= FILTER_END)].copy()
        df_wallet = df_wallet.drop_duplicates()

    # ── Core per-order reconciliation builder ──────────
    def build_reconciliation(inc_df):
        pivot = inc_df.pivot_table(index='Order Number', columns='Fee Name',
                                    values='Amount(Include Tax)', aggfunc='sum', fill_value=0)
        other_fee_names = [c for c in pivot.columns if c not in FEE_COLS + ['Item Price Credit', 'Lost Claim']]

        base = inc_df.groupby('Order Number').agg(
            order_date=('Order Creation Date', 'first'),
            order_status=('Order Status', 'first'),
            product_name=('Product Name', 'first'),
            total_released=('Amount(Include Tax)', 'sum'),
        ).reset_index()

        base = base.set_index('Order Number')
        for c in ['Item Price Credit', 'Lost Claim'] + FEE_COLS:
            base[c] = pivot[c] if c in pivot.columns else 0.0
        base['Other Fees'] = pivot[other_fee_names].sum(axis=1) if other_fee_names else 0.0
        base = base.reset_index()

        rows = []
        for _, r in base.iterrows():
            oid = r['Order Number']
            info = acc_info.get(oid)

            inv_no = info['Doc. No.'] if info is not None else ''
            inv_date = info['Date'] if info is not None else ''
            acc_subtotal = float(info['Sub-Total (ex)']) if info is not None else None
            acc_cancelled = str(info.get('Cancelled', 'F')).strip().upper() == 'T' if info is not None else False

            or_info = or_map.get(inv_no, {}) if inv_no else {}
            in_or = bool(inv_no and inv_no in or_map)
            or_paid = or_info.get('paid')
            or_outstanding = or_info.get('outstanding')

            item_price = float(r['Item Price Credit']) + float(r['Lost Claim'])
            gross_diff = round(item_price - acc_subtotal, 2) if acc_subtotal is not None else None
            price_ratio = (item_price / acc_subtotal) if acc_subtotal not in (None, 0) else None
            price_underpaid = price_ratio is not None and price_ratio < 0.995

            lost_3pl = str(r['order_status']).strip().lower() == 'lost by 3pl'
            fee_total = sum(abs(float(r[c])) for c in FEE_COLS)
            cancelled_with_fee = (acc_cancelled or lost_3pl) and fee_total > 0.01

            if cancelled_with_fee:
                issue = 'Cancelled/Lost+Fee Charged'; color = C_YELLOW
            elif not inv_no:
                issue = 'No Inv No';                  color = C_RED
            elif price_underpaid:
                issue = 'Lazada LOWER than Invoice';   color = C_PINK
            elif not in_or:
                issue = 'Not in OR (yet)';             color = C_GRAY
            elif or_outstanding and or_outstanding > 0.01:
                issue = 'OR Outstanding';              color = C_ORANGE
            else:
                issue = 'OK';                          color = C_BLUE

            underpaid_amt = round(-gross_diff, 2) if price_underpaid and gross_diff is not None else None
            real_shipfee = real_shipfee_map.get(oid)

            rows.append({
                'Order Number':        oid,
                'Inv No':              inv_no,
                'Invoice Date':        inv_date,
                'Order Creation Date': r['order_date'],
                'Order Status':        r['order_status'],
                'Product Name':        r['product_name'],
                'ACC INV Sub-Total':   acc_subtotal,
                'Income Item Price':   round(item_price, 2),
                'Lost Claim':          round(float(r['Lost Claim']), 2),
                'Real Shipping Fee(RM)': round(float(real_shipfee), 2) if real_shipfee is not None and not pd.isna(real_shipfee) else None,
                'Gross Diff(Income-Inv)': gross_diff,
                'Price Ratio(vs Inv)': round(price_ratio, 4) if price_ratio is not None else None,
                'Underpaid Amount(RM)':underpaid_amt,
                'Commission':          round(float(r['Commission']), 2),
                'Payment Fee':         round(float(r['Payment Fee']), 2),
                'LazCoins Discount':   round(float(r['LazCoins Discount']), 2),
                'LazCoins Promo Fee':  round(float(r['LazCoins Discount Promotion Fee']), 2),
                'Sponsored Affiliates':round(float(r['Sponsored Affiliates']), 2),
                'Other Fees':          round(float(r['Other Fees']), 2),
                'Total Released(Net)': round(float(r['total_released']), 2),
                'Net Revenue(excl.ShipFee)': round(float(r['total_released']) - float(real_shipfee), 2)
                    if real_shipfee is not None and not pd.isna(real_shipfee) else None,
                'OR Paid Amount':      or_paid,
                'OR Outstanding':      or_outstanding,
                'Cancelled(ACC INV)':  'Yes' if acc_cancelled else 'No',
                'Issue':               issue,
                'row_color':           color,
            })
        return pd.DataFrame(rows)

    df_main = build_reconciliation(inc)
    df_all_orders = build_reconciliation(inc_all)
    df_settlement = build_reconciliation(inc_settlement)

    # ── Sheet 5 data: invoiced but no income row this period ──
    income_orders = set(df_main['Order Number'].astype(str))
    acc_unmatched = acc_in_period[~acc_in_period['Order Number'].isin(income_orders)].copy()
    acc_unmatched = acc_unmatched[['Doc. No.', 'Date', 'Order Number', 'Sub-Total (ex)', 'Total', 'Outstanding', 'Cancelled']]
    acc_unmatched.columns = ['Inv No', 'Invoice Date', 'Order Number', 'Sub-Total', 'Total', 'Outstanding', 'Cancelled']

    # ── Sheet 4 data: cancel/lost analysis ─────────────
    df_cancel = df_main[(df_main['Cancelled(ACC INV)'] == 'Yes') |
                         (df_main['Order Status'].astype(str).str.lower() == 'lost by 3pl')].copy()
    df_cancel['Fees Charged (Total)'] = (df_cancel['Commission'].abs() + df_cancel['Payment Fee'].abs() +
                                          df_cancel['LazCoins Discount'].abs() + df_cancel['LazCoins Promo Fee'].abs() +
                                          df_cancel['Sponsored Affiliates'].abs())

    df_cancel_settlement = df_settlement[(df_settlement['Cancelled(ACC INV)'] == 'Yes') |
                                          (df_settlement['Order Status'].astype(str).str.lower() == 'lost by 3pl')].copy()
    df_cancel_settlement['Fees Charged (Total)'] = (
        df_cancel_settlement['Commission'].abs() + df_cancel_settlement['Payment Fee'].abs() +
        df_cancel_settlement['LazCoins Discount'].abs() + df_cancel_settlement['LazCoins Promo Fee'].abs() +
        df_cancel_settlement['Sponsored Affiliates'].abs())

    # ── Sheet 6 data: fee ledger totals (PDF-matching) ─
    fee_name_totals = inc_period.groupby('Fee Name')['Amount(Include Tax)'].sum()

    shipfee_pdf_scope = None
    df_shipfee_detail = None
    if real_shipfee_by_line:
        pdf_scope_lines = inc_period[inc_period['Fee Name'].isin(['Item Price Credit', 'Lost Claim'])].copy()
        pdf_scope_lines['Real Shipping Fee(RM)'] = pdf_scope_lines['Order Line ID'].map(real_shipfee_by_line)
        shipfee_pdf_scope = pdf_scope_lines['Real Shipping Fee(RM)'].sum()
        df_shipfee_detail = pdf_scope_lines[pdf_scope_lines['Real Shipping Fee(RM)'].fillna(0) != 0][
            ['Order Number', 'Order Line ID', 'Transaction Date', 'Fee Name', 'Amount(Include Tax)',
             'Seller SKU', 'Product Name', 'Real Shipping Fee(RM)']
        ].sort_values('Real Shipping Fee(RM)', ascending=False)

    known_names = {n for _, names in CATEGORY_MAP for n in names}
    leftover_names = [n for n in fee_name_totals.index if n not in known_names]
    sales_total = fee_name_totals.get('Item Price Credit', 0.0)

    # ── Summary numbers (based on settlement-date view) ─
    total_orders   = len(df_settlement)
    matched_inv    = (df_settlement['Inv No'] != '').sum()
    no_inv         = (df_settlement['Inv No'] == '').sum()
    ok_cnt         = (df_settlement['Issue'] == 'OK').sum()
    mismatch_cnt   = (df_settlement['Issue'] == 'Lazada LOWER than Invoice').sum()
    cancel_fee_cnt = (df_settlement['Issue'] == 'Cancelled/Lost+Fee Charged').sum()
    not_in_or_cnt  = (df_settlement['Issue'] == 'Not in OR (yet)').sum()
    outstanding_cnt= (df_settlement['Issue'] == 'OR Outstanding').sum()
    total_gross_diff = df_settlement['Gross Diff(Income-Inv)'].fillna(0).sum()
    total_underpaid  = df_settlement['Underpaid Amount(RM)'].fillna(0).sum()
    total_released = df_settlement['Total Released(Net)'].sum()
    total_shipfee_settlement = df_settlement['Real Shipping Fee(RM)'].fillna(0).sum()
    total_net_revenue = total_released - total_shipfee_settlement
    total_fees_charged_on_cancel = df_cancel_settlement['Fees Charged (Total)'].sum() if len(df_cancel_settlement) else 0.0

    # ═══════════════════════════════════════════════════
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────
    ws1 = wb.active
    ws1.title = '1_Summary'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 42
    ws1.column_dimensions['B'].width = 22

    ws1.merge_cells('A1:B1')
    t = ws1['A1']
    t.value = f'Lazada Reconciliation Summary  ({FILTER_START:%d %b %Y} - {FILTER_END:%d %b %Y}, by settlement date)'
    t.font = Font(bold=True, size=16, color=C_HEADER)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 36

    summary_sections = [
        ('BASIS OF THIS PAGE', '', C_HEADER, True),
        ('These numbers = Sheet 9 (by Statement/Transaction', '', C_YELLOW, False),
        ('Date - when Lazada paid you), matching Lazada\'s own', '', C_YELLOW, False),
        ('Transaction Overview export. Sheet 2 (by Order Creation', '', C_YELLOW, False),
        ('Date) is kept separately for reference, not used here.', '', C_YELLOW, False),
        ('', '', None, False),
        ('ORDER COUNTS', '', C_HEADER, True),
        ('Income Total Orders', total_orders, None, False),
        ('Matched Inv No (ACC INV)', matched_inv, C_GREEN if matched_inv == total_orders else None, False),
        ('No Inv No Found (RED)', no_inv, C_RED if no_inv > 0 else C_GREEN, False),
        ('', '', None, False),
        ('MONEY CHECK (vs your own ACC INV invoice)', '', C_HEADER, True),
        ('OK - same price or within 10% SST range (Blue)', ok_cnt, C_BLUE, False),
        ('Lazada gave LESS than your invoice - CHECK (Pink)', mismatch_cnt, C_PINK if mismatch_cnt > 0 else C_GREEN, False),
        ('Total Amount Lazada Underpaid You (RM)', round(total_underpaid, 2), C_PINK if total_underpaid > 0 else C_GREEN, False),
        ('Total Gross Diff, Income - ACC INV (RM)', round(total_gross_diff, 2), None, False),
        ('', '', None, False),
        ('NOTE ON PRICE DIFF', '', C_HEADER, True),
        ('Lazada Item Price is often ~0-10% higher than your', '', C_BLUE, False),
        ('own invoice - this matches Malaysia SST (10% tax),', '', C_BLUE, False),
        ('so it is normal and NOT flagged. Only orders where', '', C_BLUE, False),
        ('Lazada gave LESS (Pink rows) are flagged for you', '', C_BLUE, False),
        ('to check - not explainable by SST.', '', C_BLUE, False),
        ('', '', None, False),
        ('OR (COLLECTION) STATUS', '', C_HEADER, True),
        ('Not in OR yet (Gray - OR data incomplete)', not_in_or_cnt, C_GRAY, False),
        ('OR Outstanding (Orange)', outstanding_cnt, C_ORANGE if outstanding_cnt > 0 else C_GREEN, False),
        ('', '', None, False),
        ('CANCEL / LOST BY 3PL', '', C_HEADER, True),
        ('Cancelled/Lost orders still charged fee (Yellow)', cancel_fee_cnt, C_YELLOW if cancel_fee_cnt > 0 else C_GREEN, False),
        ('Total fees wrongly charged on cancel/lost (RM)', round(total_fees_charged_on_cancel, 2), C_RED if total_fees_charged_on_cancel > 0 else C_GREEN, False),
        ('', '', None, False),
        ('AMOUNT SUMMARY', '', C_HEADER, True),
        ('Total Released (Net, RM)', round(total_released, 2), None, False),
        ('Total Real Shipping Fee (RM) - not your money', round(total_shipfee_settlement, 2), C_YELLOW, False),
        ('Total Net Revenue, excl. Shipping Fee (RM)', round(total_net_revenue, 2), C_GREEN, False),
        ('', '', None, False),
        ('每张表(SHEET)是做什么的', '', C_HEADER, True),
        ('2_Reconciliation = 这个月下的单(按下单日期),订单角度', '', None, False),
        ('3_Problems = 只列有问题的单(2的筛选版)', '', None, False),
        ('4_Cancel_Lost_Analysis = 取消/搞丢但还被收费的单', '', None, False),
        ('5_Inv_No_Income_Yet = 已开发票但这次资料无收入记录', '', None, False),
        ('6_Fee_Summary_AutoCount = 费用总表,方便key in AutoCount', '', None, False),
        ('7_Wallet_CrossCheck = 钱包入账/提现明细', '', None, False),
        ('8_Full_Reconciliation_AllMonths = 全部订单,不分月份', '', None, False),
        ('9_By_SettlementDate = 这个月结算到钱的单(现金流角度,Summary用这份)', '', None, False),
        ('10_ShipFee_PDF_Detail = PDF口径运费的逐行明细', '', None, False),
    ]

    for i, (lbl, val, bg, is_sec) in enumerate(summary_sections, 3):
        ws1.row_dimensions[i].height = 22
        ca = ws1.cell(row=i, column=1, value=lbl)
        cb = ws1.cell(row=i, column=2, value=val if val != '' else '')
        if is_sec:
            for cell in (ca, cb):
                cell.fill = mk_fill(C_HEADER)
                cell.font = Font(bold=True, color=C_WHITE, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws1.merge_cells(f'A{i}:B{i}')
        else:
            ca.font = Font(size=10)
            ca.alignment = Alignment(horizontal='left', vertical='center')
            cb.font = Font(bold=True, size=11)
            cb.alignment = Alignment(horizontal='center', vertical='center')
            if bg:
                ca.fill = mk_fill(bg); cb.fill = mk_fill(bg)
            if lbl:
                ca.border = mk_border(); cb.border = mk_border()

    # ── Sheet 2: Main reconciliation ────────────────────
    ws2 = wb.create_sheet('2_Reconciliation')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A3'

    COLS2 = [
        ('Order Number', 18), ('Inv No', 16), ('Invoice Date', 13),
        ('Order Creation Date', 13), ('Order Status', 12), ('Product Name', 30),
        ('ACC INV Sub-Total', 15), ('Income Item Price', 15), ('Lost Claim', 12),
        ('Real Shipping Fee(RM)', 16), ('Gross Diff(Income-Inv)', 15),
        ('Price Ratio(vs Inv)', 14), ('Underpaid Amount(RM)', 16),
        ('Commission', 12), ('Payment Fee', 12), ('LazCoins Discount', 14),
        ('LazCoins Promo Fee', 14), ('Sponsored Affiliates', 14), ('Other Fees', 12),
        ('Total Released(Net)', 16), ('Net Revenue(excl.ShipFee)', 18),
        ('OR Paid Amount', 13), ('OR Outstanding', 13),
        ('Cancelled(ACC INV)', 13), ('Issue', 22),
    ]
    COL_ORDER2 = [c[0] for c in COLS2]
    write_note(ws2, (f'这里只列 Order Creation Date(下单日期)在 {FILTER_START:%d %b %Y}-{FILTER_END:%d %b %Y} 的单 - '
                      '如果你要的"这个月"是钱到账的时间(结算日期),不是下单日期,改看 Sheet 9。'
                      '每张单的费用是抓它"全部"的结算记录(不管哪个月结的),因为 Lazada 常常把同一张单的费用'
                      '拆到下一个月才放(比如广告费),只看当月会漏看。所以这里的金额才是这张单最终会拿到的真正总数。'
                      '"Income Item Price"已经把 Lost Claim(物流搞丢的赔偿金)加进去了,不然搞丢的单会被误判成完全没收到货款。'
                      '"Real Shipping Fee(RM)" = 从 Lazada 官方订单导出(Order Export)里的真实运费栏位算出来的 - '
                      '这笔钱是客户付的运费,借你的户口过一手,不是你的收入。"Net Revenue(excl.ShipFee)" = '
                      'Total Released(Net) 减掉 Real Shipping Fee,才是真正属于你的钱,这个才应该拿去跟你的真实收入对账。'
                      '"Underpaid Amount(RM)" 是另一个不一样的东西: 发票金额跟 Lazada 认列货款的差额,不等于运费。'
                      '颜色代表 Issue 那一栏的状态: '
                      '蓝色 OK=价钱正常/已对上; 粉红色 Lazada LOWER than Invoice=Lazada给的钱比你发票少,不是税的关系,要查; '
                      '红色 No Inv No=在你的发票登记表(ACC INV)里找不到这张单,可能还没开发票; '
                      '灰色 Not in OR (yet)=有发票但这次给的收款记录(OR)还没涵盖到这张单; '
                      '橙色 OR Outstanding=OR 里这张发票还有余额没收清; '
                      '黄色 Cancelled/Lost+Fee Charged=订单取消或被物流搞丢了,但 Lazada 还在收费,要去申诉。'), len(COL_ORDER2), height=84)
    write_header(ws2, 2, COL_ORDER2, [c[1] for c in COLS2])
    ws2.row_dimensions[2].height = 32

    AMT_COLS2 = {'ACC INV Sub-Total', 'Income Item Price', 'Lost Claim', 'Real Shipping Fee(RM)', 'Gross Diff(Income-Inv)', 'Underpaid Amount(RM)', 'Commission',
                 'Payment Fee', 'LazCoins Discount', 'LazCoins Promo Fee',
                 'Sponsored Affiliates', 'Other Fees', 'Total Released(Net)', 'Net Revenue(excl.ShipFee)',
                 'OR Paid Amount', 'OR Outstanding'}
    PCT_COLS2 = {'Price Ratio(vs Inv)'}

    write_rows(ws2, df_main, COL_ORDER2, AMT_COLS2, pct_cols=PCT_COLS2, start_row=3)

    # ── Sheet 3: Problems only ──────────────────────────
    ws3 = wb.create_sheet('3_Problems')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A3'

    df_prob = df_main[df_main['Issue'] != 'OK'].sort_values('Issue').copy()
    write_note(ws3, ('跟 Sheet 2 一模一样的栏位,但只列出有问题的单(Issue 不是 OK 的),'
                      '方便你不用自己筛。颜色/Issue 的意思看 Sheet 2 说明。'), len(COL_ORDER2))
    write_header(ws3, 2, COL_ORDER2, [c[1] for c in COLS2])
    ws3.row_dimensions[2].height = 32
    write_rows(ws3, df_prob, COL_ORDER2, AMT_COLS2, pct_cols=PCT_COLS2, start_row=3)

    # ── Sheet 4: Cancel / Lost analysis ─────────────────
    ws4 = wb.create_sheet('4_Cancel_Lost_Analysis')
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A3'

    COLS4 = [
        ('Order Number', 18), ('Inv No', 16), ('Order Status', 12),
        ('Cancelled(ACC INV)', 13), ('Commission', 12), ('Payment Fee', 12),
        ('LazCoins Discount', 14), ('LazCoins Promo Fee', 14),
        ('Sponsored Affiliates', 14), ('Fees Charged (Total)', 16),
    ]
    COL_ORDER4 = [c[0] for c in COLS4]
    write_note(ws4, ('这里列出:你的发票登记表(ACC INV)标记 Cancelled=T 的单,或者 Lazada Income '
                      '里 Order Status=lost by 3pl(被物流搞丢)的单。"Fees Charged (Total)"是这张单'
                      '取消/搞丢之后 Lazada 还收了多少手续费/佣金 - 这笔钱理论上应该要退回给你,'
                      '如果没退,可以拿这份去跟 Lazada 申诉。'), len(COL_ORDER4))
    write_header(ws4, 2, COL_ORDER4, [c[1] for c in COLS4])
    ws4.row_dimensions[2].height = 28
    AMT_COLS4 = {'Commission', 'Payment Fee', 'LazCoins Discount', 'LazCoins Promo Fee',
                 'Sponsored Affiliates', 'Fees Charged (Total)'}

    for r_i, (_, row) in enumerate(df_cancel.iterrows(), 3):
        for c_i, col in enumerate(COL_ORDER4, 1):
            val = row.get(col, '')
            if isinstance(val, float) and pd.isna(val):
                val = ''
            cell = ws4.cell(row=r_i, column=c_i, value=val)
            cell.fill = mk_fill(C_YELLOW); cell.border = mk_border()
            if col in AMT_COLS4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    if len(df_cancel) > 0:
        tr = len(df_cancel) + 3
        ws4.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True)
        cell = ws4.cell(row=tr, column=COL_ORDER4.index('Fees Charged (Total)') + 1,
                         value=round(df_cancel['Fees Charged (Total)'].sum(), 2))
        cell.number_format = '#,##0.00'; cell.font = Font(bold=True)
        cell.fill = mk_fill('FFEB3B')

    # ── Sheet 5: invoices with no matching Income row this period ──
    ws5 = wb.create_sheet('5_Inv_No_Income_Yet')
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = 'A3'

    ws5.merge_cells('A1:G1')
    note5 = ws5['A1']
    note5.value = ('这些单已经开了发票(有 Inv No),但这次的 Lazada Income Overview 资料里,'
                   '找不到这张单的收入记录 - 很可能是这张单不在这次的对账月份内(比如发票开得比较早/比较晚),'
                   '不是发票号有问题')
    note5.font = Font(italic=True, size=9, color='555555')
    note5.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws5.row_dimensions[1].height = 30

    COLS5 = [('Inv No', 16), ('Invoice Date', 13), ('Order Number', 18),
             ('Sub-Total', 13), ('Total', 13), ('Outstanding', 13), ('Cancelled', 11)]
    COL_ORDER5 = [c[0] for c in COLS5]
    write_header(ws5, 2, COL_ORDER5, [c[1] for c in COLS5])
    ws5.row_dimensions[2].height = 28
    AMT_COLS5 = {'Sub-Total', 'Total', 'Outstanding'}

    for r_i, (_, row) in enumerate(acc_unmatched.iterrows(), 3):
        for c_i, col in enumerate(COL_ORDER5, 1):
            val = row.get(col, '')
            if isinstance(val, float) and pd.isna(val):
                val = ''
            cell = ws5.cell(row=r_i, column=c_i, value=val)
            cell.fill = mk_fill(C_GRAY); cell.border = mk_border()
            if col in AMT_COLS5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # ── Sheet 6: Fee Summary for AutoCount keyin ────────
    ws6 = wb.create_sheet('6_Fee_Summary_AutoCount')
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions['A'].width = 22
    ws6.column_dimensions['B'].width = 42
    ws6.column_dimensions['C'].width = 16
    ws6.column_dimensions['D'].width = 12

    ws6.merge_cells('A1:D1')
    note6 = ws6['A1']
    note6.value = ('这一页是"这个月 Lazada 实际放了多少钱到我账户"的总表 - 按 Transaction Date(结算日期)'
                   '筛选这个月的资料,不管订单是哪个月下的 - 这个跟 Lazada 官方 PDF 月结单的算法一样,'
                   '所以数字应该完全对得上 PDF,可以逐行拿去核对、直接 key in AutoCount。'
                   '(注意跟 Sheet 2 不一样: Sheet 2 是"这个月的订单,最终总共赚多少",这一页是"这个月总共收到多少钱",'
                   '两个概念不同,金额不会一样。) % 那一栏 = 这个费用 / 总货款(Sales) 的比例。'
                   '最下面"Real Shipping Fee"那一行,是照 PDF 的 Sales+Claims 范围去 Order Export 查真实运费 - '
                   '这笔钱不是你的钱,是客户运费借你户口过一手。')
    note6.font = Font(italic=True, size=9, color='555555')
    note6.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws6.row_dimensions[1].height = 74

    write_header(ws6, 2, ['Category', 'Fee Name (raw, from Income Overview)', 'Amount (RM)', '% of Sales'],
                 [22, 42, 16, 12])
    ws6.row_dimensions[2].height = 26

    r_i = 3
    AMT_FMT = '#,##0.00'
    grand_total = 0.0
    category_totals = {}
    for cat, names in CATEGORY_MAP:
        cat_total = 0.0
        if names:
            for name in names:
                amt = float(fee_name_totals.get(name, 0.0))
                cat_total += amt
                pct = (amt / sales_total) if sales_total else 0.0
                ca = ws6.cell(row=r_i, column=1, value='')
                cb = ws6.cell(row=r_i, column=2, value=name)
                cc = ws6.cell(row=r_i, column=3, value=round(amt, 2))
                cd = ws6.cell(row=r_i, column=4, value=pct)
                for c in (ca, cb, cc, cd):
                    c.border = mk_border()
                cc.number_format = AMT_FMT; cc.alignment = Alignment(horizontal='right')
                cd.number_format = '0.0%'; cd.alignment = Alignment(horizontal='right')
                cb.alignment = Alignment(horizontal='left')
                r_i += 1
        else:
            ca = ws6.cell(row=r_i, column=1, value='')
            cb = ws6.cell(row=r_i, column=2, value='(none this period - normal, PDF also shows 0.00)')
            cc = ws6.cell(row=r_i, column=3, value=0.0)
            cd = ws6.cell(row=r_i, column=4, value=0.0)
            for c in (ca, cb, cc, cd):
                c.border = mk_border()
            cc.number_format = AMT_FMT; cc.alignment = Alignment(horizontal='right')
            cd.number_format = '0.0%'; cd.alignment = Alignment(horizontal='right')
            cb.font = Font(italic=True, color='888888')
            r_i += 1

        pct_total = (cat_total / sales_total) if sales_total else 0.0
        ca = ws6.cell(row=r_i, column=1, value=cat)
        cb = ws6.cell(row=r_i, column=2, value='SUBTOTAL')
        cc = ws6.cell(row=r_i, column=3, value=round(cat_total, 2))
        cd = ws6.cell(row=r_i, column=4, value=pct_total)
        for c in (ca, cb, cc, cd):
            c.font = Font(bold=True)
            c.fill = mk_fill(C_HEADER if cat == 'Sales' else C_GRAY)
            c.border = mk_border()
        if cat == 'Sales':
            for c in (ca, cb, cc, cd):
                c.font = Font(bold=True, color=C_WHITE)
        cc.number_format = AMT_FMT; cc.alignment = Alignment(horizontal='right')
        cd.number_format = '0.0%'; cd.alignment = Alignment(horizontal='right')
        ca.alignment = Alignment(horizontal='left'); cb.alignment = Alignment(horizontal='left')
        r_i += 1
        grand_total += cat_total if cat != 'Sales' else 0.0
        category_totals[cat] = cat_total

    if leftover_names:
        for name in leftover_names:
            amt = float(fee_name_totals.get(name, 0.0))
            grand_total += amt
            pct = (amt / sales_total) if sales_total else 0.0
            ca = ws6.cell(row=r_i, column=1, value='Other (uncategorised)')
            cb = ws6.cell(row=r_i, column=2, value=name)
            cc = ws6.cell(row=r_i, column=3, value=round(amt, 2))
            cd = ws6.cell(row=r_i, column=4, value=pct)
            for c in (ca, cb, cc, cd):
                c.border = mk_border(); c.fill = mk_fill(C_YELLOW)
            cc.number_format = AMT_FMT; cc.alignment = Alignment(horizontal='right')
            cd.number_format = '0.0%'; cd.alignment = Alignment(horizontal='right')
            r_i += 1

    r_i += 1
    grand_total_with_sales = sales_total + grand_total
    ca = ws6.cell(row=r_i, column=1, value='TOTAL SETTLEMENT')
    cb = ws6.cell(row=r_i, column=2, value='Sales + all fees above (should = Total Released Net in Sheet 2)')
    cc = ws6.cell(row=r_i, column=3, value=round(grand_total_with_sales, 2))
    cd = ws6.cell(row=r_i, column=4, value=1.0 if sales_total else 0.0)
    for c in (ca, cb, cc, cd):
        c.font = Font(bold=True); c.fill = mk_fill(C_GREEN); c.border = mk_border()
    cc.number_format = AMT_FMT; cc.alignment = Alignment(horizontal='right')
    cd.number_format = '0.0%'; cd.alignment = Alignment(horizontal='right')
    ca.alignment = Alignment(horizontal='left'); cb.alignment = Alignment(horizontal='left')

    if shipfee_pdf_scope is not None:
        r_i += 1
        pct_ship = (shipfee_pdf_scope / sales_total) if sales_total else 0.0
        ca = ws6.cell(row=r_i, column=1, value='Real Shipping Fee (not your money)')
        cb = ws6.cell(row=r_i, column=2, value='From Order Export, same Sales+Claims lines as this PDF period')
        cc = ws6.cell(row=r_i, column=3, value=round(float(shipfee_pdf_scope), 2))
        cd = ws6.cell(row=r_i, column=4, value=pct_ship)
        for c in (ca, cb, cc, cd):
            c.font = Font(bold=True); c.fill = mk_fill(C_YELLOW); c.border = mk_border()
        cc.number_format = AMT_FMT; cc.alignment = Alignment(horizontal='right')
        cd.number_format = '0.0%'; cd.alignment = Alignment(horizontal='right')
        ca.alignment = Alignment(horizontal='left'); cb.alignment = Alignment(horizontal='left')

    # ── Sheet 7: Wallet cross-check ──────────────────────
    if len(df_wallet):
        ws7 = wb.create_sheet('7_Wallet_CrossCheck')
        ws7.sheet_view.showGridLines = False
        ws7.freeze_panes = 'A3'
        COLS7 = [('Transaction Time', 16), ('Type', 12), ('Sub Type', 16),
                 ('Statement No', 20), ('Amount', 14), ('Remarks', 38)]
        COL_ORDER7 = [c[0] for c in COLS7]
        write_note(ws7, ('钱包(Balance)的每一笔入账/提现记录,不是订单明细。蓝色 Deposit=Lazada 把这个'
                          'Statement 结算的钱放进你的钱包; 橙色 Withdrawal=钱包自动提现去银行。'
                          'Statement No 那栏可以拿去跟 Sheet 6 的总数对,确认每个结算周期的钱有没有全部入账。'), len(COL_ORDER7))
        write_header(ws7, 2, COL_ORDER7, [c[1] for c in COLS7])
        ws7.row_dimensions[2].height = 26
        for r_i, (_, row) in enumerate(df_wallet.iterrows(), 3):
            clr = C_BLUE if row['Type'] == 'Deposit' else C_ORANGE
            for c_i, col in enumerate(COL_ORDER7, 1):
                val = row.get(col, '')
                if isinstance(val, float) and pd.isna(val):
                    val = ''
                cell = ws7.cell(row=r_i, column=c_i, value=val)
                cell.fill = mk_fill(clr); cell.border = mk_border()
                if col == 'Amount':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    # ── Sheet 8: Full reconciliation, every order in the source file ──
    ws8 = wb.create_sheet('8_Full_Reconciliation_AllMonths')
    ws8.sheet_view.showGridLines = False
    ws8.freeze_panes = 'A3'

    write_note(ws8, ('这份是从原始、还没整理过的资料(Income Overview + ACC INV + Order Export)独立重新算出来的,'
                      f'涵盖全部 {df_all_orders["Order Number"].nunique()} 张单,不分月份 - '
                      '栏位跟 Sheet 2 一模一样。"Real Shipping Fee(RM)" 是从 Lazada Order Export 抓的真实运费。'
                      '"Underpaid Amount(RM)" 是另一个概念(发票 vs 货款差额),不是运费,两个分开看。'),
              len(COL_ORDER2), height=56)
    write_header(ws8, 2, COL_ORDER2, [c[1] for c in COLS2])
    ws8.row_dimensions[2].height = 32
    df_all_orders_sorted = df_all_orders.sort_values('Order Creation Date')
    write_rows(ws8, df_all_orders_sorted, COL_ORDER2, AMT_COLS2, pct_cols=PCT_COLS2, start_row=3)

    # ── Sheet 9: Reconciliation by Settlement Date ──────
    ws9 = wb.create_sheet('9_By_SettlementDate')
    ws9.sheet_view.showGridLines = False
    ws9.freeze_panes = 'A3'

    write_note(ws9, (f'这里用"钱什么时候结算给你"(Transaction Date/Statement Period 在 '
                       f'{FILTER_START:%d %b %Y}-{FILTER_END:%d %b %Y})来抓单,不是用下单日期 - '
                       '所以像那种上个月下单、但这个月才结算给钱的单,会出现在这里而不是 Sheet 2。'
                       '跟 Sheet 2 一样,一旦这张单符合条件,会抓它"全部"的费用记录来算,不会漏看分开结算的部分。'
                       '栏位跟 Sheet 2 一模一样。如果你要的"这个月"是指现金流(钱到账时间),用这份;'
                       '如果是指订单本身(哪个月卖出),用 Sheet 2。'),
              len(COL_ORDER2), height=68)
    write_header(ws9, 2, COL_ORDER2, [c[1] for c in COLS2])
    ws9.row_dimensions[2].height = 32
    df_settlement_sorted = df_settlement.sort_values('Order Creation Date')
    write_rows(ws9, df_settlement_sorted, COL_ORDER2, AMT_COLS2, pct_cols=PCT_COLS2, start_row=3)

    # ── Sheet 10: line-by-line detail behind the PDF-scope Shipping Fee ──
    if df_shipfee_detail is not None and len(df_shipfee_detail):
        ws10 = wb.create_sheet('10_ShipFee_PDF_Detail')
        ws10.sheet_view.showGridLines = False
        ws10.freeze_panes = 'A3'

        COLS10 = [('Order Number', 18), ('Order Line ID', 18), ('Transaction Date', 13),
                  ('Fee Name', 16), ('Amount(Include Tax)', 16), ('Seller SKU', 11),
                  ('Product Name', 34), ('Real Shipping Fee(RM)', 16)]
        COL_ORDER10 = [c[0] for c in COLS10]
        write_note(ws10, (f'这是 Sheet 6 那行 "Real Shipping Fee (not your money)" = {shipfee_pdf_scope:,.2f} '
                           f'的完整明细,一共 {len(df_shipfee_detail)} 行 - 每一行是一个 Order Line,'
                           'Fee Name 是 Item Price Credit 或 Lost Claim,Transaction Date 在这个月的 PDF 范围内,'
                           '"Real Shipping Fee(RM)" 是从 Order Export 查到的真实运费,全部加起来就是上面那个总数。'),
                  len(COL_ORDER10), height=48)
        write_header(ws10, 2, COL_ORDER10, [c[1] for c in COLS10])
        ws10.row_dimensions[2].height = 28
        AMT_COLS10 = {'Amount(Include Tax)', 'Real Shipping Fee(RM)'}

        for r_i, (_, row) in enumerate(df_shipfee_detail.iterrows(), 3):
            for c_i, col in enumerate(COL_ORDER10, 1):
                val = row.get(col, '')
                if isinstance(val, float) and pd.isna(val):
                    val = ''
                cell = ws10.cell(row=r_i, column=c_i, value=val)
                cell.fill = mk_fill(C_YELLOW); cell.border = mk_border()
                if col in AMT_COLS10:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col in ('Order Number', 'Order Line ID'):
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        tr = len(df_shipfee_detail) + 3
        ws10.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True)
        tot_cell = ws10.cell(row=tr, column=COL_ORDER10.index('Real Shipping Fee(RM)') + 1,
                              value=round(float(shipfee_pdf_scope), 2))
        tot_cell.number_format = '#,##0.00'; tot_cell.font = Font(bold=True)
        tot_cell.fill = mk_fill(C_GREEN)

    # ── Save to bytes ────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stats = {
        'total_orders': int(total_orders),
        'matched_inv': int(matched_inv),
        'no_inv': int(no_inv),
        'ok_cnt': int(ok_cnt),
        'mismatch_cnt': int(mismatch_cnt),
        'not_in_or_cnt': int(not_in_or_cnt),
        'outstanding_cnt': int(outstanding_cnt),
        'cancel_fee_cnt': int(cancel_fee_cnt),
        'total_gross_diff': float(total_gross_diff),
        'total_underpaid': float(total_underpaid),
        'total_released': float(total_released),
        'total_shipfee_settlement': float(total_shipfee_settlement),
        'total_net_revenue': float(total_net_revenue),
        'total_fees_charged_on_cancel': float(total_fees_charged_on_cancel),
        'df_prob': df_prob,
        # PDF-matching (Sheet 6) figures - this is the ledger scoped exactly
        # like the official Lazada PDF statement, for cross-checking against it.
        'pdf_sales': float(sales_total),
        'pdf_category_totals': {k: float(v) for k, v in category_totals.items()},
        'pdf_total_settlement': float(grand_total_with_sales),
    }
    return buf, stats


def parse_pdf_statement(pdf_file):
    """Extract the Income summary category totals from a Lazada 'My Income
    Statement' PDF. Returns a dict with sales/lazada_fees/logistics/
    marketing_fees/claims/total_settlement, or None if the expected line
    isn't found (e.g. wrong PDF, or a layout Lazada has changed)."""
    import re
    import pypdf
    reader = pypdf.PdfReader(pdf_file)
    text = '\n'.join((p.extract_text() or '') for p in reader.pages)

    m = re.search(
        r'Settlement\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+'
        r'(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})', text)
    if not m:
        return None
    nums = [float(g.replace(',', '')) for g in m.groups()]
    result = {
        'sales': nums[0], 'lazada_fees': nums[1], 'logistics': nums[2],
        'marketing_fees': nums[3], 'claims': nums[4], 'total_settlement': nums[5],
    }
    period = re.search(r'Statement from\s+(\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})', text)
    if period:
        result['period_start'] = period.group(1)
        result['period_end'] = period.group(2)
    return result
