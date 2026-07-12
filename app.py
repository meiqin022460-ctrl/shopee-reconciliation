import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, warnings, os
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Shopee 对账系统", page_icon="📊", layout="centered")

# ── Password protection ─────────────────────────────────
def check_password():
    correct = st.secrets.get("APP_PASSWORD", "shopee2024")
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    st.title("📊 Shopee 对账系统")
    st.markdown("---")
    pw = st.text_input("请输入密码：", type="password")
    if st.button("登录", type="primary"):
        if pw == correct:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("密码错误，请重试。")
    return False

if not check_password():
    st.stop()
# ────────────────────────────────────────────────────────

st.title("📊 Shopee 对账系统")
st.markdown("---")

# ── Shared helpers ─────────────────────────────────────
def mk_fill(h):
    return PatternFill(start_color=h, end_color=h, fill_type='solid')

def mk_border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row_num, labels, widths=None, bg='2F4F8F'):
    f = mk_fill(bg)
    ft = Font(bold=True, color='FFFFFF', size=10)
    al = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, lbl in enumerate(labels, 1):
        c = ws.cell(row=row_num, column=i, value=lbl)
        c.fill = f; c.font = ft; c.alignment = al; c.border = mk_border()
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def safe_num(s):
    return pd.to_numeric(s, errors='coerce').fillna(0)

# ══════════════════════════════════════════════════════
# TWO TABS
# ══════════════════════════════════════════════════════
tab1, tab2 = st.tabs(["① 加 Inv No 到 Income 文件", "② 完整对账报告（含 OR）"])


# ╔════════════════════════════════════════════════════╗
# ║  TAB 1 — Add Inv No to Income                     ║
# ╚════════════════════════════════════════════════════╝
with tab1:
    st.subheader("把 ACC INV 的 Doc. No. 加到 Income 报表")
    st.caption("上传 ACC INV 和 Income 文件 → 系统自动匹配 Shipping Info = Order ID → 在 Income 右边加一列 'Inv No' → 下载")
    st.markdown("")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**① ACC INV 发票文件**")
        t1_5151 = st.file_uploader("ACC INV", type=["xlsx","xls"], key="t1_5151",
                                    label_visibility="collapsed")
        if t1_5151: st.success(f"✅ {t1_5151.name}")

    with c2:
        st.markdown("**② Income 报表**")
        t1_income = st.file_uploader("Income", type=["xlsx","xls"], key="t1_income",
                                      label_visibility="collapsed")
        if t1_income: st.success(f"✅ {t1_income.name}")

    st.markdown("")

    if t1_5151 and t1_income:
        st.success("两个文件都好了，点下面开始！")

        if st.button("🚀  加入 Inv No 列", type="primary", use_container_width=True, key="btn1"):
            with st.spinner("处理中..."):

                # Load 5151
                df5 = pd.read_excel(t1_5151, sheet_name='Sheet', header=0,
                                    dtype={'Doc. No.': str, 'Shipping Info': str})
                df5.columns = [str(c).strip() for c in df5.columns]
                ship_to_inv = dict(zip(df5['Shipping Info'].astype(str),
                                       df5['Doc. No.'].astype(str)))

                # Load Income with pandas (fast — find header row first)
                t1_income.seek(0)
                raw = pd.read_excel(t1_income, sheet_name='Income', header=None)
                hdr_idx = 2  # default
                for i in range(min(10, len(raw))):
                    if 'Order ID' in raw.iloc[i].astype(str).values:
                        hdr_idx = i; break
                t1_income.seek(0)
                df_inc2 = pd.read_excel(t1_income, sheet_name='Income', header=hdr_idx,
                                        dtype={'Order ID': str})
                df_inc2.columns = [str(c).strip() for c in df_inc2.columns]

                if 'Order ID' not in df_inc2.columns:
                    st.error("找不到 'Order ID' 列，请确认文件格式正确。")
                    st.stop()

                # Keep only Order rows (drop SKU rows)
                if 'View By' in df_inc2.columns:
                    df_inc2 = df_inc2[df_inc2['View By'].astype(str).str.strip() == 'Order'].reset_index(drop=True)

                # Add Inv No column right after Order ID
                inv_vals, red_flag = [], []
                for oid in df_inc2['Order ID'].astype(str):
                    inv = ship_to_inv.get(oid.strip(), '')
                    if inv and inv != 'nan':
                        inv_vals.append(inv); red_flag.append(False)
                    else:
                        inv_vals.append('❌ 找不到'); red_flag.append(True)

                matched   = sum(1 for f in red_flag if not f)
                unmatched = sum(1 for f in red_flag if f)

                oid_pos = df_inc2.columns.tolist().index('Order ID')
                cols_before = df_inc2.columns.tolist()[:oid_pos+1]
                cols_after  = df_inc2.columns.tolist()[oid_pos+1:]
                df_inc2.insert(oid_pos + 1, 'Inv No', inv_vals)

                # Write to Excel with minimal formatting (fast)
                wb2 = Workbook()
                ws2 = wb2.active; ws2.title = 'Income'
                ws2.freeze_panes = 'A2'
                col_order_out = cols_before + ['Inv No'] + cols_after

                # Header row
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                red_fill    = PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid')
                hdr_fill    = PatternFill(start_color='2F4F8F', end_color='2F4F8F', fill_type='solid')
                hdr_font    = Font(bold=True, color='FFFFFF')
                for c_i, col in enumerate(col_order_out, 1):
                    cell = ws2.cell(row=1, column=c_i, value=col)
                    if col == 'Inv No':
                        cell.fill = yellow_fill; cell.font = Font(bold=True)
                    else:
                        cell.fill = hdr_fill; cell.font = hdr_font
                ws2.column_dimensions[get_column_letter(col_order_out.index('Inv No')+1)].width = 18
                ws2.column_dimensions['A'].width = 22

                # Data rows
                for r_i, (_, row) in enumerate(df_inc2[col_order_out].iterrows(), 2):
                    is_red = red_flag[r_i - 2]
                    for c_i, col in enumerate(col_order_out, 1):
                        val = row[col]
                        if isinstance(val, float) and pd.isna(val): val = ''
                        cell = ws2.cell(row=r_i, column=c_i, value=val)
                        if col == 'Inv No':
                            cell.fill = red_fill if is_red else yellow_fill
                            if is_red: cell.font = Font(color='CC0000')

                buf = io.BytesIO()
                wb2.save(buf)
                buf.seek(0)

            # Results
            total = matched + unmatched
            st.success(f"✅ 完成！{matched:,} 行匹配到 Inv No，{unmatched:,} 行找不到")

            if unmatched > 0:
                pct = unmatched / total * 100
                st.warning(f"⚠️ {unmatched} 行（{pct:.1f}%）在 5151 找不到对应的 Shipping Info — 可能是系统漏开单，或订单不属于这个店铺。")

            # Suggest output filename
            orig_name = t1_income.name.replace('.xlsx','').replace('.xls','')
            out_name = f"{orig_name}_with_InvNo.xlsx"

            st.download_button(
                label="📥  下载加了 Inv No 的 Income 文件",
                data=buf,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
    else:
        missing = []
        if not t1_5151:  missing.append("ACC INV 文件")
        if not t1_income: missing.append("Income 文件")
        st.info(f"还需要上传：{' / '.join(missing)}")


# ╔════════════════════════════════════════════════════╗
# ║  TAB 2 — Full Reconciliation                      ║
# ╚════════════════════════════════════════════════════╝
with tab2:
    st.subheader("完整对账（ACC INV + Income + OR）")
    st.caption("上传文件 → 生成彩色对账报告，包含退货、遗失赔偿、取消单等全部情况")
    st.markdown("")

    ca, cb, cc = st.columns(3)
    with ca:
        st.markdown("**① ACC INV 发票文件**（可多选）")
        t2_5151 = st.file_uploader("ACC INV", type=["xlsx","xls"], key="t2_5151",
                                    accept_multiple_files=True, label_visibility="collapsed")
        if t2_5151:
            for f in t2_5151: st.success(f"✅ {f.name}")
    with cb:
        st.markdown("**② Income 报表**（可多选）")
        t2_income = st.file_uploader("Income", type=["xlsx","xls"], key="t2_income",
                                      accept_multiple_files=True, label_visibility="collapsed")
        if t2_income:
            for f in t2_income: st.success(f"✅ {f.name}")
    with cc:
        st.markdown("**③ OR 收款文件**（可多选）")
        t2_or = st.file_uploader("OR", type=["xlsx","xls"], key="t2_or",
                                  accept_multiple_files=True, label_visibility="collapsed")
        if t2_or:
            for f in t2_or: st.success(f"✅ {f.name}")

    st.markdown("")
    cd, ce = st.columns(2)
    with cd:
        st.markdown("**④ Order.all 全部订单**（选填，可多选）")
        st.caption("用来显示取消单、退货单、下个月才到款的单")
        t2_orders = st.file_uploader("Order All", type=["xlsx","xls"], key="t2_orders",
                                      accept_multiple_files=True, label_visibility="collapsed")
        if t2_orders:
            for f in t2_orders: st.success(f"✅ {f.name}")
    with ce:
        st.markdown("**⑤ Wallet Balance 报告**（选填，可多选）")
        st.caption("用来显示包裹遗失赔偿、物流赔偿")
        t2_wallet = st.file_uploader("Wallet", type=["xlsx","xls"], key="t2_wallet",
                                      accept_multiple_files=True, label_visibility="collapsed")
        if t2_wallet:
            for f in t2_wallet: st.success(f"✅ {f.name}")

    st.markdown("")

    all_3 = t2_5151 and t2_income and t2_or
    if not all_3:
        missing = []
        if not t2_5151:  missing.append("ACC INV 文件")
        if not t2_income: missing.append("Income 文件")
        if not t2_or:    missing.append("OR 收款文件")
        st.info(f"还需要上传：{' / '.join(missing)}")
    else:
        total_files = len(t2_5151) + len(t2_income) + len(t2_or)
        st.success(f"共 {total_files} 个文件，准备好了！")

        if st.button("🚀  开始完整对账", type="primary", use_container_width=True, key="btn2"):
            with st.spinner("匹配中，请稍等..."):

                C_BLUE   = 'ADD8E6'
                C_ORANGE = 'FFD580'
                C_RED    = 'FFB3B3'
                C_YELLOW = 'FFFAAA'
                C_GRAY   = 'D3D3D3'
                C_GREEN  = 'C8E6C9'
                C_HEADER = '2F4F8F'
                C_WHITE  = 'FFFFFF'
                PCT_COLS = {'Commission %','Service Fee %','Txn Fee %','Total Fee %'}
                AMT_COLS = {'Product Price (RM)','Net Payout (RM)','OR Org Amt','OR Pay','Amt Diff(Price-OR)',
                            'Commission','Service Fee','Txn Fee','Platform Fees Total',
                            'Voucher(Seller)','AMS Fee','Ads Escrow','Calc Subtotal','Payout Gap',
                            'Shipping(Buyer)','Refund Amount','OR Outstanding'}

                # Load files — merge multiple uploads into one dataframe each

                # 5151: stack all uploaded files
                df5_list = []
                for f in t2_5151:
                    _d = pd.read_excel(f, sheet_name='Sheet', header=0,
                                       dtype={'Doc. No.': str, 'Shipping Info': str})
                    _d.columns = [str(c).strip() for c in _d.columns]
                    df5_list.append(_d)
                df5 = pd.concat(df5_list, ignore_index=True).drop_duplicates(subset=['Doc. No.'])

                # Income: stack all uploaded files
                inc_list = []
                for f in t2_income:
                    raw = pd.read_excel(f, sheet_name='Income', header=None)
                    hdr_row = 2
                    for i in range(min(10, len(raw))):
                        if 'Order ID' in raw.iloc[i].astype(str).values:
                            hdr_row = i; break
                    f.seek(0)
                    _d = pd.read_excel(f, sheet_name='Income', header=hdr_row, dtype={'Order ID': str})
                    _d.columns = [str(c).strip() for c in _d.columns]
                    if 'View By' in _d.columns:
                        _d = _d[_d['View By'].astype(str).str.strip() == 'Order']
                    inc_list.append(_d)
                df_inc = pd.concat(inc_list, ignore_index=True).drop_duplicates(subset=['Order ID'])

                ship_to_inv = dict(zip(df5['Shipping Info'].astype(str),
                                       df5['Doc. No.'].astype(str)))
                inv_info    = df5.set_index('Doc. No.').to_dict('index')

                # OR: load each file separately, track file source
                # Color palette per OR file (up to 6 files)
                OR_FILE_COLORS = [
                    'AED6F1',  # OR1 蓝
                    'A9DFBF',  # OR2 绿
                    'D7BDE2',  # OR3 紫
                    'FAD7A0',  # OR4 橙黄
                    'F9EBEA',  # OR5 浅粉
                    'B2EBF2',  # OR6 青
                    'FCF3CF',  # OR7 淡黄
                    'D5F5E3',  # OR8 薄荷绿
                    'FADBD8',  # OR9 玫瑰
                    'D6EAF8',  # OR10 天蓝
                    'E8DAEF',  # OR11 薰衣草
                    'FDEBD0',  # OR12 米黄
                ]
                OR_FILE_NAMES  = [f.name for f in t2_or]

                or_map = {}
                for file_idx, f in enumerate(t2_or):
                    _d = pd.read_excel(f, sheet_name='Sheet', header=0, dtype={'No.': str})
                    _d.columns = [str(c).strip() for c in _d.columns]
                    or_cols = list(_d.columns)
                    pay_col_f       = 'Pay' if 'Pay' in or_cols else ('Paid Amount' if 'Paid Amount' in or_cols else None)
                    knock_off_col_f = 'Knock Off Date' if 'Knock Off Date' in or_cols else None
                    select_col_f    = 'Select' if 'Select' in or_cols else None

                    for _, row in _d.iterrows():
                        no = str(row['No.']).strip()
                        if no in or_map:
                            continue  # already seen in earlier file
                        pay_val         = float(row.get(pay_col_f, 0) or 0) if pay_col_f else 0.0
                        outstanding_val = float(row.get('Outstanding', 0) or 0)
                        knock_off_val   = str(row.get(knock_off_col_f, '') or '').strip() if knock_off_col_f else ''
                        select_val      = str(row.get(select_col_f, '') or '').strip().lower() if select_col_f else ''
                        collected = (select_val == 'checked') or \
                                    (pay_val > 0 and outstanding_val == 0) or \
                                    (knock_off_val not in ('', 'nan', 'None', 'NaT'))
                        or_map[no] = {
                            'org_amt':     float(row.get('Org. Amt.', 0) or 0),
                            'pay':         pay_val,
                            'outstanding': outstanding_val,
                            'collected':   collected,
                            'date':        row.get('Date',''),
                            'knock_off':   knock_off_val,
                            'file_idx':    file_idx,
                            'file_name':   OR_FILE_NAMES[file_idx],
                        }

                # ── Order.all: load all orders (optional) ──────────
                ord_all = {}  # order_id → {status, cancel_reason, return_status, complete_time, grand_total}
                if t2_orders:
                    ord_list = []
                    for f in t2_orders:
                        _d = pd.read_excel(f, sheet_name='orders', header=0, dtype={'Order ID': str})
                        _d.columns = [str(c).strip() for c in _d.columns]
                        ord_list.append(_d)
                    df_ord = pd.concat(ord_list, ignore_index=True).drop_duplicates(subset=['Order ID'])
                    for _, row in df_ord.iterrows():
                        oid = str(row['Order ID']).strip()
                        ord_all[oid] = {
                            'status':        str(row.get('Order Status', '') or '').strip(),
                            'cancel_reason': str(row.get('Cancel reason', '') or '').strip(),
                            'return_status': str(row.get('Return / Refund Status', '') or '').strip(),
                            'complete_time': row.get('Order Complete Time', ''),
                            'grand_total':   float(row.get('Grand Total', 0) or 0),
                        }

                # ── Wallet: load balance transactions (optional) ────
                wallet_map = {}  # order_id → list of {type, description, direction, amount}
                if t2_wallet:
                    for f in t2_wallet:
                        _raw = pd.read_excel(f, sheet_name='Transaction Report', header=None)
                        # find header row (row with 'Transaction Type')
                        w_hdr = 17
                        for i in range(30):
                            vals = [str(v) for v in _raw.iloc[i].tolist()]
                            if 'Transaction Type' in vals:
                                w_hdr = i; break
                        _d = pd.read_excel(f, sheet_name='Transaction Report', header=w_hdr, dtype={'Order ID': str})
                        _d.columns = [str(c).strip() for c in _d.columns]
                        _d = _d.dropna(subset=['Transaction Type'])
                        for _, row in _d.iterrows():
                            oid = str(row.get('Order ID', '') or '').strip()
                            if not oid or oid == 'nan': continue
                            entry = {
                                'type':        str(row.get('Transaction Type', '') or ''),
                                'description': str(row.get('Description', '') or ''),
                                'direction':   str(row.get('Money Direction', '') or ''),
                                'amount':      float(row.get('Amount', 0) or 0) if 'Amount' in _d.columns else 0.0,
                            }
                            wallet_map.setdefault(oid, []).append(entry)

                num_map = {
                    'Total Released Amount (RM)':              'total_released',
                    'Product Price':                           'product_price',
                    'Refund Amount':                           'refund_amt',
                    'Commission Fee (incl. SST)':              'commission',
                    'Service Fee (Incl. SST)':                 'service_fee',
                    'Transaction Fee (Incl. SST)':             'txn_fee',
                    'AMS Commission Fee':                      'ams_fee',
                    'Ads Escrow Top Up Fee':                   'ads_escrow',
                    'Voucher Sponsored by Seller':             'voucher_seller',
                    'Shipping Fee Paid by Buyer (excl. SST)':  'shipping_buyer',
                    'Shipping Fee Charged by Logistic Provider':'shipping_logistic',
                    'Shipping Rebate From Shopee':             'shipping_rebate',
                }
                for col, alias in num_map.items():
                    df_inc[alias] = safe_num(df_inc[col]) if col in df_inc.columns else 0.0

                agg = df_inc.groupby('Order ID').agg(
                    order_date       =('Order Creation Date',  'first'),
                    payout_date      =('Payout Completed Date','first'),
                    total_released   =('total_released',  'sum'),
                    product_price    =('product_price',   'sum'),
                    refund_amt       =('refund_amt',      'sum'),
                    commission       =('commission',      'sum'),
                    service_fee      =('service_fee',     'sum'),
                    txn_fee          =('txn_fee',         'sum'),
                    ams_fee          =('ams_fee',         'sum'),
                    ads_escrow       =('ads_escrow',      'sum'),
                    voucher_seller   =('voucher_seller',  'sum'),
                    shipping_buyer   =('shipping_buyer',  'sum'),
                    shipping_logistic=('shipping_logistic','sum'),
                    shipping_rebate  =('shipping_rebate', 'sum'),
                    item_count       =('total_released',  'count'),
                ).reset_index()
                agg['oid_str'] = agg['Order ID'].astype(str)

                rows_main = []
                for _, r in agg.iterrows():
                    oid    = r['oid_str']
                    inv_no = ship_to_inv.get(oid, '')

                    has_refund     = float(r['refund_amt']) < 0
                    cancelled_5151 = False
                    inv_date       = ''

                    if inv_no and inv_no in inv_info:
                        info = inv_info[inv_no]
                        cancelled_5151 = str(info.get('Cancelled','F')).strip().upper() == 'T'
                        inv_date = info.get('Date','')

                    in_or        = bool(inv_no and inv_no in or_map)
                    od           = or_map.get(inv_no, {}) if inv_no else {}
                    or_org_amt   = od.get('org_amt')
                    or_pay       = od.get('pay')
                    or_outstanding = od.get('outstanding')
                    or_collected = od.get('collected', False)
                    or_knock_off = od.get('knock_off','')
                    or_file_idx  = od.get('file_idx', None)
                    or_file_name = od.get('file_name', '')

                    # Order.all info
                    ord_info      = ord_all.get(oid, {})
                    ord_status    = ord_info.get('status', '')
                    ord_return    = ord_info.get('return_status', '')

                    # Wallet info for this order
                    w_entries     = wallet_map.get(oid, [])
                    w_notes       = '; '.join(set(e['description'] for e in w_entries if 'Income' not in e['type'])) if w_entries else ''

                    prod_price     = round(float(r['product_price']), 2)
                    net_payout     = round(float(r['total_released']), 2)
                    voucher_seller = round(float(r['voucher_seller']), 2)
                    ams_fee        = round(float(r['ams_fee']), 2)
                    ads_escrow     = round(float(r['ads_escrow']), 2)
                    amt_diff       = round(prod_price - or_org_amt, 2) if or_org_amt is not None else None
                    total_fees     = abs(float(r['commission'])) + abs(float(r['service_fee'])) + abs(float(r['txn_fee'])) + abs(ams_fee) + abs(ads_escrow)
                    # Subtotal check: Product Price + Voucher + Commission + Service + Txn + AMS + Ads Escrow
                    calc_subtotal  = round(prod_price + voucher_seller + float(r['commission']) +
                                          float(r['service_fee']) + float(r['txn_fee']) + ams_fee + ads_escrow, 2)
                    payout_gap     = round(calc_subtotal - net_payout, 2)
                    cancel_fee = (has_refund or cancelled_5151) and total_fees > 0.01

                    base = prod_price if prod_price != 0 else None
                    def pct(v): return round(abs(v) / base * 100, 2) if base else None

                    if cancel_fee:
                        issue = 'Cancelled+Fee'; color = C_YELLOW
                    elif not inv_no:
                        issue = 'No Inv No';     color = C_RED
                    elif not in_or:
                        issue = 'Not in OR';     color = C_ORANGE
                    else:
                        # Color by which OR file matched (regardless of collected/outstanding)
                        file_color = OR_FILE_COLORS[or_file_idx % len(OR_FILE_COLORS)] if or_file_idx is not None else C_BLUE
                        if or_collected:
                            issue = 'OK'
                        else:
                            issue = 'OR Outstanding'
                        color = file_color

                    rows_main.append({
                        'Order ID':              oid,
                        'Inv No':                inv_no,
                        'Invoice Date':          inv_date,
                        'Payout Date':           r['payout_date'],
                        'Items':                 int(r['item_count']),
                        'Product Price (RM)':    prod_price,
                        'Net Payout (RM)':       net_payout,
                        'OR Org Amt':            or_org_amt,
                        'OR Pay':                or_pay,
                        'Amt Diff(Price-OR)':    amt_diff,
                        'Commission':            round(float(r['commission']), 2),
                        'Commission %':          pct(float(r['commission'])),
                        'Service Fee':           round(float(r['service_fee']), 2),
                        'Service Fee %':         pct(float(r['service_fee'])),
                        'Txn Fee':               round(float(r['txn_fee']), 2),
                        'Txn Fee %':             pct(float(r['txn_fee'])),
                        'Platform Fees Total':   round(-total_fees, 2),
                        'Total Fee %':           round(total_fees / base * 100, 2) if base else None,
                        'Voucher(Seller)':       voucher_seller,
                        'AMS Fee':               ams_fee,
                        'Ads Escrow':            ads_escrow,
                        'Calc Subtotal':         calc_subtotal,
                        'Payout Gap':            payout_gap,
                        'Shipping(Buyer)':       round(float(r['shipping_buyer']), 2),
                        'Refund Amount':         round(float(r['refund_amt']), 2),
                        'OR Outstanding':        or_outstanding,
                        'OR Knock Off Date':     or_knock_off,
                        'OR File':               or_file_name,
                        'Cancelled(ACC INV)':    'Yes' if cancelled_5151 else 'No',
                        'Order Status':          ord_status,
                        'Return/Refund Status':  ord_return,
                        'Wallet Note':           w_notes,
                        'Issue':                 issue,
                        'row_color':             color,
                    })

                df_main = pd.DataFrame(rows_main)
                # Sort: group by OR file first, then by issue within each group
                # OR-matched rows: sort by file_idx; unmatched at the bottom
                def or_file_sort(row):
                    color = row.get('row_color','')
                    for i, c in enumerate(OR_FILE_COLORS):
                        if color == c:
                            return i  # OR file 1=0, 2=1, 3=2...
                    return 90  # Not in OR / No Inv / Cancel → at the bottom
                issue_order = {'OK':0,'OR Outstanding':1,'Not in OR':2,'Cancelled+Fee':3,'No Inv No':4}
                df_main['_file_sort']  = df_main.apply(or_file_sort, axis=1)
                df_main['_issue_sort'] = df_main['Issue'].map(issue_order).fillna(9)
                df_main = df_main.sort_values(['_file_sort','_issue_sort']).drop(columns=['_file_sort','_issue_sort']).reset_index(drop=True)

                income_invs  = set(df_main['Inv No'].astype(str))
                or_unmatched = [
                    {'Inv No':no,'OR File':od.get('file_name',''),'OR Date':od['date'],
                     'OR Org Amt':od['org_amt'],
                     'OR Pay':od['pay'],'OR Outstanding':od['outstanding'],
                     'Collected':'Yes' if od['collected'] else 'No',
                     'Knock Off Date':od['knock_off'],
                     'row_color': OR_FILE_COLORS[od.get('file_idx',0) % len(OR_FILE_COLORS)],
                     'Note':'In OR but no matching Order ID in Income (probably prior month)'}
                    for no, od in or_map.items() if no not in income_invs
                ]
                df_or_unmatched = pd.DataFrame(or_unmatched) if or_unmatched else pd.DataFrame()

                df_cancel = df_main[(df_main['Refund Amount'] < 0) | (df_main['Cancelled(ACC INV)'] == 'Yes')].copy()
                df_cancel['Platform Fee Charged'] = df_cancel['Platform Fees Total'].abs()
                df_cancel['Refund Covered?'] = df_cancel.apply(
                    lambda x: 'Yes' if abs(x['Refund Amount']) >= x['Platform Fee Charged'] else 'No', axis=1)
                df_cancel['Potential Claim (RM)'] = df_cancel.apply(
                    lambda x: round(x['Platform Fee Charged'] - abs(x['Refund Amount']), 2)
                              if x['Refund Covered?'] == 'No' else 0.0, axis=1)

                total_orders  = len(df_main)
                matched_inv   = (df_main['Inv No'] != '').sum()
                no_inv        = (df_main['Inv No'] == '').sum()
                collected     = (df_main['Issue'] == 'OK').sum()
                outstanding   = df_main['Issue'].isin(['OR Outstanding','Not in OR']).sum()
                cancel_cnt    = len(df_cancel)
                total_claim   = df_cancel['Potential Claim (RM)'].sum() if len(df_cancel) else 0.0
                total_released= df_main['Product Price (RM)'].sum()
                total_or_pay  = df_main['OR Pay'].fillna(0).sum()
                df_prob       = df_main[~df_main['Issue'].isin(['OK', 'Cancelled+Fee'])]

                # Build Excel
                COLS2 = [
                    ('Order ID',22),('Inv No',18),('Invoice Date',14),('Payout Date',14),
                    ('Items',7),('Product Price (RM)',18),('Net Payout (RM)',16),
                    ('OR Org Amt',13),('OR Pay',11),
                    ('Amt Diff(Price-OR)',16),
                    ('Commission',13),('Commission %',11),
                    ('Service Fee',12),('Service Fee %',11),
                    ('Txn Fee',11),('Txn Fee %',11),
                    ('AMS Fee',11),('Ads Escrow',13),('Voucher(Seller)',14),
                    ('Platform Fees Total',16),('Total Fee %',11),
                    ('Calc Subtotal',16),('Payout Gap',14),
                    ('Shipping(Buyer)',14),
                    ('Refund Amount',13),('OR Outstanding',14),('OR Knock Off Date',14),
                    ('OR File',22),('Cancelled(ACC INV)',12),
                    ('Order Status',14),('Return/Refund Status',18),('Wallet Note',35),('Issue',15),
                ]
                COL_ORDER2 = [c[0] for c in COLS2]

                def write_rows(ws, df, col_order):
                    for r_i, (_, row) in enumerate(df.iterrows(), 2):
                        clr = row.get('row_color', 'FFFFFF')
                        for c_i, col in enumerate(col_order, 1):
                            val = row.get(col, '')
                            if isinstance(val, float) and pd.isna(val): val = ''
                            cell = ws.cell(row=r_i, column=c_i, value=val)
                            # Payout Gap cell: red if gap != 0, green if 0
                            if col == 'Payout Gap':
                                gap = val if isinstance(val, (int, float)) else 0
                                if abs(gap) > 0.02:
                                    cell.fill = mk_fill('FF6B6B')  # Red
                                    cell.font = Font(bold=True, color='FFFFFF')
                                else:
                                    cell.fill = mk_fill('C8E6C9')  # Green
                            else:
                                cell.fill = mk_fill(clr)
                            cell.border = mk_border()
                            if col in AMT_COLS:
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif col in PCT_COLS:
                                cell.number_format = '0.00"%"'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif col in ('Order ID','Inv No'):
                                cell.number_format = '@'
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='center')

                wb2 = Workbook()

                # Sheet 1: Summary
                ws1 = wb2.active; ws1.title = '1_Summary'
                ws1.sheet_view.showGridLines = False
                ws1.column_dimensions['A'].width = 42
                ws1.column_dimensions['B'].width = 22
                ws1.merge_cells('A1:B1')
                t = ws1['A1']
                t.value = 'Shopee Reconciliation Summary'
                t.font = Font(bold=True, size=16, color=C_HEADER)
                t.alignment = Alignment(horizontal='center', vertical='center')
                ws1.row_dimensions[1].height = 36

                avg_fee_pct = df_main['Total Fee %'].dropna().mean() if 'Total Fee %' in df_main.columns else 0

                summary_rows = [
                    ('ORDER COUNTS','',C_HEADER,True),
                    ('Income Total Orders', total_orders, None, False),
                    ('Matched Inv No', matched_inv, C_GREEN if matched_inv==total_orders else None, False),
                    ('No Inv No (RED)', no_inv, C_RED if no_inv>0 else C_GREEN, False),
                    ('','',None,False),
                    ('PAYMENT STATUS','',C_HEADER,True),
                    ('Collected  [BLUE]', collected, C_BLUE, False),
                    ('Outstanding / Not in OR  [ORANGE]', outstanding, C_ORANGE if outstanding>0 else C_GREEN, False),
                    ('','',None,False),
                    ('FEES','',C_HEADER,True),
                    ('Avg Platform Fee %', f'{avg_fee_pct:.2f}%', None, False),
                    ('','',None,False),
                    ('CANCEL / RETURNS','',C_HEADER,True),
                    ('Cancel/Refund Orders  [YELLOW]', cancel_cnt, C_YELLOW if cancel_cnt>0 else C_GREEN, False),
                    ('Potential Claim (RM)', round(total_claim,2), C_RED if total_claim>0 else C_GREEN, False),
                    ('','',None,False),
                    ('AMOUNT','',C_HEADER,True),
                    ('Total Product Price - Income (RM)', round(total_released,2), None, False),
                    ('Total OR Pay Collected (RM)', round(total_or_pay,2), None, False),
                ]
                for i,(lbl,val,bg,is_sec) in enumerate(summary_rows, 3):
                    ws1.row_dimensions[i].height = 22
                    ca2 = ws1.cell(row=i, column=1, value=lbl)
                    cb2 = ws1.cell(row=i, column=2, value=val if val!='' else '')
                    if is_sec:
                        for cell in (ca2,cb2):
                            cell.fill=mk_fill(C_HEADER); cell.font=Font(bold=True,color='FFFFFF',size=10)
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                        ws1.merge_cells(f'A{i}:B{i}')
                    else:
                        ca2.font=Font(size=10); ca2.alignment=Alignment(horizontal='left',vertical='center')
                        cb2.font=Font(bold=True,size=11); cb2.alignment=Alignment(horizontal='center',vertical='center')
                        if bg: ca2.fill=mk_fill(bg); cb2.fill=mk_fill(bg)
                        if lbl: ca2.border=mk_border(); cb2.border=mk_border()

                ws2b = wb2.create_sheet('2_Reconciliation')
                ws2b.sheet_view.showGridLines = False

                # --- Color legend rows at top ---
                legend_items = []
                for i, fname in enumerate(OR_FILE_NAMES):
                    legend_items.append((f'OR{i+1}: {fname}', OR_FILE_COLORS[i % len(OR_FILE_COLORS)]))
                legend_items.append(('Not in OR（找不到）', C_ORANGE))
                legend_items.append(('No Inv No（无发票）', C_RED))
                legend_items.append(('Cancelled+Fee（退单仍扣费）', C_YELLOW))

                leg_row = 1
                for lbl, clr in legend_items:
                    cell = ws2b.cell(row=leg_row, column=1, value=lbl)
                    cell.fill = mk_fill(clr)
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = mk_border()
                    ws2b.row_dimensions[leg_row].height = 18
                    leg_row += 1

                # blank separator
                ws2b.row_dimensions[leg_row].height = 8
                leg_row += 1

                hdr_row_xl = leg_row
                write_header(ws2b, hdr_row_xl, COL_ORDER2, [c[1] for c in COLS2])
                ws2b.row_dimensions[hdr_row_xl].height = 32
                ws2b.freeze_panes = ws2b.cell(row=hdr_row_xl + 1, column=1).coordinate

                # Override write_rows to start after legend
                def write_rows_offset(ws, df, col_order, start_row):
                    for r_i, (_, row) in enumerate(df.iterrows(), start_row):
                        clr = row.get('row_color', 'FFFFFF')
                        for c_i, col in enumerate(col_order, 1):
                            val = row.get(col, '')
                            if isinstance(val, float) and pd.isna(val): val = ''
                            cell = ws.cell(row=r_i, column=c_i, value=val)
                            if col == 'Payout Gap':
                                gap = val if isinstance(val, (int, float)) else 0
                                if abs(gap) > 0.02:
                                    cell.fill = mk_fill('FF6B6B')
                                    cell.font = Font(bold=True, color='FFFFFF')
                                else:
                                    cell.fill = mk_fill('C8E6C9')
                            elif col == 'Amt Diff(Price-OR)':
                                diff = val if isinstance(val, (int, float)) else None
                                if diff is not None and abs(diff) > 0.02:
                                    cell.fill = mk_fill('FF6B6B')
                                    cell.font = Font(bold=True, color='FFFFFF')
                                else:
                                    cell.fill = mk_fill(clr)
                            else:
                                cell.fill = mk_fill(clr)
                            cell.border = mk_border()
                            if col in AMT_COLS:
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif col in PCT_COLS:
                                cell.number_format = '0.00"%"'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif col in ('Order ID','Inv No'):
                                cell.number_format = '@'
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='center')

                write_rows_offset(ws2b, df_main, COL_ORDER2, hdr_row_xl + 1)

                ws3 = wb2.create_sheet('3_Problems')
                ws3.sheet_view.showGridLines = False
                # Color legend (same as Sheet 2)
                leg3 = 1
                for lbl, clr in legend_items:
                    cell = ws3.cell(row=leg3, column=1, value=lbl)
                    cell.fill = mk_fill(clr)
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = mk_border()
                    ws3.row_dimensions[leg3].height = 18
                    leg3 += 1
                ws3.row_dimensions[leg3].height = 8
                leg3 += 1
                hdr3 = leg3
                write_header(ws3, hdr3, COL_ORDER2, [c[1] for c in COLS2])
                ws3.row_dimensions[hdr3].height = 32
                ws3.freeze_panes = ws3.cell(row=hdr3 + 1, column=1).coordinate
                write_rows_offset(ws3, df_prob.sort_values('Issue'), COL_ORDER2, hdr3 + 1)

                ws4 = wb2.create_sheet('4_Cancel_Analysis')
                ws4.sheet_view.showGridLines = False; ws4.freeze_panes = 'A2'
                COLS4 = [('Order ID',22),('Inv No',18),('Cancelled(ACC INV)',13),
                         ('Refund Amount',14),('Commission',13),('Service Fee',12),
                         ('Txn Fee',11),('Platform Fee Charged',18),
                         ('Refund Covered?',14),('Potential Claim (RM)',18),('Issue',15)]
                COL4 = [c[0] for c in COLS4]
                write_header(ws4, 1, COL4, [c[1] for c in COLS4])
                AMT4 = {'Refund Amount','Commission','Service Fee','Txn Fee','Platform Fee Charged','Potential Claim (RM)'}
                for r_i,(_, row) in enumerate(df_cancel.iterrows(), 2):
                    for c_i, col in enumerate(COL4, 1):
                        val = row.get(col,'')
                        if isinstance(val, float) and pd.isna(val): val = ''
                        cell = ws4.cell(row=r_i, column=c_i, value=val)
                        cell.fill = mk_fill('FFFAAA'); cell.border = mk_border()
                        if col in AMT4:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                if len(df_cancel) > 0:
                    tr = len(df_cancel) + 2
                    ws4.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True)
                    for c_i, col in enumerate(COL4, 1):
                        cell = ws4.cell(row=tr, column=c_i)
                        cell.fill = mk_fill('FFEB3B'); cell.border = mk_border()
                        if col == 'Platform Fee Charged':
                            cell.value = round(df_cancel['Platform Fee Charged'].sum(),2)
                            cell.number_format='#,##0.00'; cell.font=Font(bold=True)
                        elif col == 'Potential Claim (RM)':
                            cell.value = round(df_cancel['Potential Claim (RM)'].sum(),2)
                            cell.number_format='#,##0.00'; cell.font=Font(bold=True)

                ws5 = wb2.create_sheet('5_OR_Unmatched')
                ws5.sheet_view.showGridLines = False
                COLS5=[('Inv No',18),('OR File',28),('OR Date',14),('OR Org Amt',13),('OR Pay',11),
                       ('OR Outstanding',14),('Collected',10),('Knock Off Date',14),('Note',45)]
                COL5=[c[0] for c in COLS5]
                AMT5={'OR Org Amt','OR Pay','OR Outstanding'}
                # Color legend at top
                leg5 = 1
                for lbl, clr in legend_items:
                    cell = ws5.cell(row=leg5, column=1, value=lbl)
                    cell.fill = mk_fill(clr)
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = mk_border()
                    ws5.row_dimensions[leg5].height = 18
                    leg5 += 1
                ws5.row_dimensions[leg5].height = 8
                leg5 += 1
                hdr5 = leg5
                write_header(ws5, hdr5, COL5, [c[1] for c in COLS5])
                ws5.row_dimensions[hdr5].height = 32
                ws5.freeze_panes = ws5.cell(row=hdr5 + 1, column=1).coordinate
                if len(df_or_unmatched)>0:
                    # Sort by OR file
                    df_or_unmatched_sorted = df_or_unmatched.sort_values('OR File') if 'OR File' in df_or_unmatched.columns else df_or_unmatched
                    for r_i,(_, row) in enumerate(df_or_unmatched_sorted.iterrows(), hdr5 + 1):
                        clr = row.get('row_color', 'D3D3D3')
                        for c_i,col in enumerate(COL5,1):
                            val=row.get(col,'')
                            if isinstance(val,float) and pd.isna(val): val=''
                            cell=ws5.cell(row=r_i,column=c_i,value=val)
                            cell.fill=mk_fill(clr); cell.border=mk_border()
                            if col in AMT5:
                                cell.number_format='#,##0.00'
                                cell.alignment=Alignment(horizontal='right',vertical='center')
                            else:
                                cell.alignment=Alignment(horizontal='left',vertical='center')

                # ── Sheet 6: Detailed Summary ──────────────────────────────
                ws6 = wb2.create_sheet('6_Detail_Summary')
                ws6.sheet_view.showGridLines = False
                ws6.column_dimensions['A'].width = 30
                ws6.column_dimensions['B'].width = 18
                ws6.column_dimensions['C'].width = 18
                ws6.column_dimensions['D'].width = 18
                ws6.column_dimensions['E'].width = 22

                def s6_title(row, text):
                    c = ws6.cell(row=row, column=1, value=text)
                    c.fill = mk_fill(C_HEADER)
                    c.font = Font(bold=True, color='FFFFFF', size=11)
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    ws6.merge_cells(f'A{row}:E{row}')
                    ws6.row_dimensions[row].height = 24

                def s6_hdr(row, labels):
                    for ci, lbl in enumerate(labels, 1):
                        c = ws6.cell(row=row, column=ci, value=lbl)
                        c.fill = mk_fill('4A7DB8')
                        c.font = Font(bold=True, color='FFFFFF', size=10)
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        c.border = mk_border()
                    ws6.row_dimensions[row].height = 20

                def s6_row(row, vals, clr='FFFFFF', bold=False):
                    for ci, val in enumerate(vals, 1):
                        c = ws6.cell(row=row, column=ci, value=val)
                        c.fill = mk_fill(clr)
                        c.font = Font(bold=bold, size=10)
                        c.border = mk_border()
                        if ci > 1 and isinstance(val, (int, float)):
                            c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            c.alignment = Alignment(horizontal='left', vertical='center')
                    ws6.row_dimensions[row].height = 18

                cur = 1

                # ── Section 1: Per OR File Match Summary ──
                s6_title(cur, '① 各 OR 文件匹配汇总'); cur += 1
                s6_hdr(cur, ['OR 文件', '匹配张数', '总 OR Org Amt (RM)', '总 OR Pay (RM)', '已收款张数']); cur += 1
                df_matched = df_main[df_main['OR File'] != '']
                for fname in OR_FILE_NAMES:
                    sub = df_matched[df_matched['OR File'] == fname]
                    collected_cnt = (sub['Issue'] == 'OK').sum()
                    clr = OR_FILE_COLORS[OR_FILE_NAMES.index(fname) % len(OR_FILE_COLORS)]
                    s6_row(cur, [fname, len(sub),
                                 round(sub['OR Org Amt'].fillna(0).sum(), 2),
                                 round(sub['OR Pay'].fillna(0).sum(), 2),
                                 collected_cnt], clr=clr); cur += 1
                # Total row
                s6_row(cur, ['TOTAL', len(df_matched),
                             round(df_matched['OR Org Amt'].fillna(0).sum(), 2),
                             round(df_matched['OR Pay'].fillna(0).sum(), 2),
                             (df_matched['Issue'] == 'OK').sum()],
                       clr='D5D8DC', bold=True); cur += 1
                cur += 1  # blank

                # ── Section 2: Amt Diff(Price-OR) Summary ──
                s6_title(cur, '② 价钱差异汇总  Amt Diff (Product Price − OR Org Amt)'); cur += 1
                df_diff = df_main[df_main['Amt Diff(Price-OR)'].apply(
                    lambda x: isinstance(x, (int, float)) and not pd.isna(x) and abs(x) > 0.02)].copy()
                s6_hdr(cur, ['Order ID', 'Inv No', 'Product Price (RM)', 'OR Org Amt (RM)', 'Amt Diff (RM)']); cur += 1
                if len(df_diff) > 0:
                    for _, rw in df_diff.iterrows():
                        clr = rw.get('row_color', 'FFFFFF')
                        s6_row(cur, [rw['Order ID'], rw['Inv No'],
                                     rw['Product Price (RM)'], rw['OR Org Amt'],
                                     rw['Amt Diff(Price-OR)']], clr=clr); cur += 1
                    s6_row(cur, ['TOTAL 差异', '', '', '',
                                 round(df_diff['Amt Diff(Price-OR)'].sum(), 2)],
                           clr='D5D8DC', bold=True); cur += 1
                else:
                    s6_row(cur, ['✅ 全部价钱一致，无差异', '', '', '', ''], clr='C8E6C9'); cur += 1
                cur += 1  # blank

                # ── Section 3: Payout Gap Summary ──
                s6_title(cur, '③ 到手金额差异汇总  Payout Gap (Calc Subtotal − Net Payout)'); cur += 1
                df_gap = df_main[df_main['Payout Gap'].apply(
                    lambda x: isinstance(x, (int, float)) and not pd.isna(x) and abs(x) > 0.02)].copy()
                s6_hdr(cur, ['Order ID', 'Inv No', 'Calc Subtotal (RM)', 'Net Payout (RM)', 'Payout Gap (RM)']); cur += 1
                if len(df_gap) > 0:
                    for _, rw in df_gap.sort_values('Payout Gap').iterrows():
                        clr = rw.get('row_color', 'FFFFFF')
                        s6_row(cur, [rw['Order ID'], rw['Inv No'],
                                     rw['Calc Subtotal'], rw['Net Payout (RM)'],
                                     rw['Payout Gap']], clr=clr); cur += 1
                    s6_row(cur, ['TOTAL 差异', '', '', '',
                                 round(df_gap['Payout Gap'].sum(), 2)],
                           clr='D5D8DC', bold=True); cur += 1
                else:
                    s6_row(cur, ['✅ 全部到手金额一致，无差异', '', '', '', ''], clr='C8E6C9'); cur += 1
                cur += 1  # blank

                # ── Section 4: Shopee Fee Breakdown ──
                ws6.column_dimensions['C'].width = 20
                ws6.column_dimensions['D'].width = 20
                ws6.column_dimensions['E'].width = 20
                s6_title(cur, '④ Shopee 平台费用明细汇总'); cur += 1
                s6_hdr(cur, ['费用类型', '总金额 (RM)', '占总货价 %', '张数 (有收费)', '平均每单 (RM)']); cur += 1

                total_price_base = df_main['Product Price (RM)'].sum()
                def fee_pct(amt):
                    return round(abs(amt) / total_price_base * 100, 2) if total_price_base else 0
                def fee_avg(amt, cnt):
                    return round(abs(amt) / cnt, 2) if cnt else 0

                fee_rows = [
                    ('佣金 Commission Fee',      df_main['Commission'].sum()),
                    ('服务费 Service Fee',        df_main['Service Fee'].sum()),
                    ('交易费 Transaction Fee',    df_main['Txn Fee'].sum()),
                    ('AMS 佣金 AMS Fee',          df_main['AMS Fee'].sum()),
                    ('广告托管 Ads Escrow',        df_main['Ads Escrow'].sum()),
                    ('卖家自付优惠券 Voucher(Seller)', df_main['Voucher(Seller)'].sum()),
                ]
                clrs = ['FDECEA','FFF3E0','E8F5E9','E3F2FD','F3E5F5','FFF9C4']
                total_all_fees = 0
                for idx, (name, total_amt) in enumerate(fee_rows):
                    cnt = (df_main['Commission'].abs() > 0.01).sum() if '佣金' in name else \
                          (df_main['Service Fee'].abs() > 0.01).sum() if '服务费' in name else \
                          (df_main['Txn Fee'].abs() > 0.01).sum() if '交易费' in name else \
                          (df_main['AMS Fee'].abs() > 0.01).sum() if 'AMS' in name else \
                          (df_main['Ads Escrow'].abs() > 0.01).sum() if '广告' in name else \
                          (df_main['Voucher(Seller)'].abs() > 0.01).sum()
                    s6_row(cur, [name, round(total_amt, 2), f'{fee_pct(total_amt):.2f}%',
                                 int(cnt), fee_avg(total_amt, cnt)], clr=clrs[idx % len(clrs)]); cur += 1
                    if '优惠券' not in name:
                        total_all_fees += abs(total_amt)

                # Grand total (excluding voucher which is seller-initiated)
                s6_row(cur, ['★ 平台抽费合计 (不含优惠券)',
                             round(-total_all_fees, 2),
                             f'{fee_pct(total_all_fees):.2f}%', '', ''],
                       clr='FFCDD2', bold=True); cur += 1
                s6_row(cur, ['总货价 (Product Price)',
                             round(total_price_base, 2), '100.00%', '', ''],
                       clr='D5D8DC', bold=True); cur += 1
                s6_row(cur, ['实际到手 (Net Payout)',
                             round(df_main['Net Payout (RM)'].sum(), 2),
                             f'{fee_pct(df_main["Net Payout (RM)"].sum()):.2f}%', '', ''],
                       clr='C8E6C9', bold=True); cur += 1
                # ────────────────────────────────────────────────────────────

                # ── Sheet 7: Special Cases ──────────────────────────────────
                if ord_all or wallet_map:
                    ws7 = wb2.create_sheet('7_Special_Cases')
                    ws7.sheet_view.showGridLines = False
                    ws7.column_dimensions['A'].width = 24
                    ws7.column_dimensions['B'].width = 20
                    ws7.column_dimensions['C'].width = 18
                    ws7.column_dimensions['D'].width = 20
                    ws7.column_dimensions['E'].width = 16
                    ws7.column_dimensions['F'].width = 45

                    def s7_title(row, text, clr):
                        c = ws7.cell(row=row, column=1, value=text)
                        c.fill = mk_fill(clr); c.font = Font(bold=True, color='FFFFFF', size=11)
                        c.alignment = Alignment(horizontal='left', vertical='center')
                        ws7.merge_cells(f'A{row}:F{row}')
                        ws7.row_dimensions[row].height = 24

                    def s7_hdr(row, labels):
                        for ci, lbl in enumerate(labels, 1):
                            c = ws7.cell(row=row, column=ci, value=lbl)
                            c.fill = mk_fill('4A7DB8'); c.font = Font(bold=True, color='FFFFFF', size=10)
                            c.alignment = Alignment(horizontal='center', vertical='center')
                            c.border = mk_border()
                        ws7.row_dimensions[row].height = 20

                    def s7_row(row, vals, clr='FFFFFF'):
                        for ci, val in enumerate(vals, 1):
                            c = ws7.cell(row=row, column=ci, value=val)
                            c.fill = mk_fill(clr); c.border = mk_border()
                            c.font = Font(size=10)
                            if ci > 1 and isinstance(val, (int, float)):
                                c.number_format = '#,##0.00'
                                c.alignment = Alignment(horizontal='right', vertical='center')
                            else:
                                c.alignment = Alignment(horizontal='left', vertical='center')
                        ws7.row_dimensions[row].height = 18

                    r7 = 1
                    inc_ids_set = set(df_main['Order ID'].astype(str))

                    # Section A: Cancelled before shipment
                    if ord_all:
                        cancelled_rows = [(oid, info) for oid, info in ord_all.items()
                                         if info['status'] == 'Cancelled' and oid not in inc_ids_set]
                        s7_title(r7, '① 出货前取消（Income 不会出现，正常）', '7F8C8D'); r7 += 1
                        s7_hdr(r7, ['Order ID', 'Inv No', 'Cancel Reason', 'Grand Total (RM)', '', '']); r7 += 1
                        for oid, info in cancelled_rows[:200]:
                            inv = ship_to_inv.get(oid, '')
                            inv = inv if inv and inv != 'nan' else '❌ 找不到'
                            s7_row(r7, [oid, inv, info['cancel_reason'][:60], info['grand_total'], '', ''], clr='E8E8E8'); r7 += 1
                        s7_row(r7, [f'共 {len(cancelled_rows)} 张取消单', '', '', '', '', ''], clr='D5D8DC'); r7 += 2

                    # Section B: Completed but payout next month
                    if ord_all:
                        next_month = [(oid, info) for oid, info in ord_all.items()
                                      if info['status'] == 'Completed' and oid not in inc_ids_set]
                        s7_title(r7, '② 已完成、下个月才到款（正常，会在下期 Income 出现）', '1A5276'); r7 += 1
                        s7_hdr(r7, ['Order ID', 'Inv No', 'Order Complete Time', 'Grand Total (RM)', '', '']); r7 += 1
                        for oid, info in next_month[:200]:
                            inv = ship_to_inv.get(oid, '')
                            inv = inv if inv and inv != 'nan' else '❌ 找不到'
                            s7_row(r7, [oid, inv, str(info['complete_time']), info['grand_total'], '', ''], clr='D6EAF8'); r7 += 1
                        s7_row(r7, [f'共 {len(next_month)} 张，预计收款 RM {sum(i["grand_total"] for _,i in next_month):,.2f}', '', '', '', '', ''], clr='AED6F1'); r7 += 2

                    # Section C: Wallet special transactions
                    if wallet_map:
                        special_wallet = []
                        for oid, entries in wallet_map.items():
                            for e in entries:
                                if 'Order Income' not in e['type']:
                                    special_wallet.append((oid, e))
                        if special_wallet:
                            s7_title(r7, '③ Wallet 特殊记录（包裹遗失赔偿、物流赔偿、退款扣回）', 'C0392B'); r7 += 1
                            s7_hdr(r7, ['Order ID', 'Transaction Type', 'Money Direction', 'Amount (RM)', '', 'Description']); r7 += 1
                            for oid, e in special_wallet:
                                clr = 'FADBD8' if 'Out' in e['direction'] else 'D5F5E3'
                                s7_row(r7, [oid, e['type'], e['direction'], e['amount'], '', e['description'][:60]], clr=clr); r7 += 1
                            s7_row(r7, [f'共 {len(special_wallet)} 笔特殊钱包记录', '', '', '', '', ''], clr='D5D8DC'); r7 += 1
                # ────────────────────────────────────────────────────────────

                buf2 = io.BytesIO()
                wb2.save(buf2); buf2.seek(0)

            # Display results
            st.success("✅ 对账完成！")
            st.markdown("---")
            st.subheader("📈 结果总览")

            r1c1,r1c2,r1c3,r1c4 = st.columns(4)
            r1c1.metric("总订单", total_orders)
            r1c2.metric("🔵 已收款", collected)
            r1c3.metric("🟠 未收款", outstanding)
            r1c4.metric("🔴 找不到 Inv", no_inv)

            r2c1,r2c2,r2c3,r2c4 = st.columns(4)
            r2c1.metric("🟡 Cancel 单", cancel_cnt)
            r2c2.metric("可追讨 (RM)", f"{total_claim:.2f}")
            r2c3.metric("平均平台费率", f"{avg_fee_pct:.2f}%")
            r2c4.metric("OR 已收 (RM)", f"{total_or_pay:,.2f}")

            st.caption("💡 Product Price (RM) = 顾客付的货价（对应 OR Org Amt）｜Net Payout (RM) = Shopee 扣费后实际打给你的钱")

            # OR File color legend
            if len(OR_FILE_NAMES) > 1:
                st.markdown("**🎨 OR 文件颜色对照：**")
                legend_cols = st.columns(len(OR_FILE_NAMES) + 2)
                palette_hex = ['#AED6F1','#A9DFBF','#D7BDE2','#FAD7A0','#FDEBD0','#B2EBF2']
                for i, fname in enumerate(OR_FILE_NAMES):
                    legend_cols[i].markdown(
                        f'<div style="background:{palette_hex[i%len(palette_hex)]};padding:6px;border-radius:4px;text-align:center;font-size:12px"><b>OR{i+1}</b><br>{fname}</div>',
                        unsafe_allow_html=True)
                legend_cols[len(OR_FILE_NAMES)].markdown('<div style="background:#FFD580;padding:6px;border-radius:4px;text-align:center;font-size:12px"><b>Not in OR</b><br>橙色</div>', unsafe_allow_html=True)
                legend_cols[len(OR_FILE_NAMES)+1].markdown('<div style="background:#FFB3B3;padding:6px;border-radius:4px;text-align:center;font-size:12px"><b>No Inv No</b><br>红色</div>', unsafe_allow_html=True)

            if len(df_prob) > 0:
                st.markdown("---")
                st.subheader(f"⚠️ 问题清单（{len(df_prob)} 行）")
                ic = df_prob['Issue'].value_counts().reset_index()
                ic.columns = ['问题类型','数量']
                st.dataframe(ic, use_container_width=True, hide_index=True)

            # OR entries with no matching Income order
            if len(df_or_unmatched) > 0:
                st.markdown("---")
                st.subheader(f"📋 OR 里有、但 Income 没有的单（{len(df_or_unmatched)} 笔）")
                st.caption("这些发票在 OR 收款系统里有记录，但这个月的 Income 报表里找不到对应的订单。可能是上个月的单，或日期范围不在这份报表内。")
                show_cols = ['Inv No','OR Date','OR Org Amt','OR Pay','OR Outstanding','Collected','Knock Off Date']
                show_cols = [c for c in show_cols if c in df_or_unmatched.columns]
                st.dataframe(
                    df_or_unmatched[show_cols].reset_index(drop=True),
                    use_container_width=True,
                    hide_index=True
                )

            st.markdown("---")
            st.download_button(
                label="📥  下载完整对账报告 Excel",
                data=buf2,
                file_name="Reconciliation_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
